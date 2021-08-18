import openpyxl
import os
from inkontor_resources.classes.product import Product
from inkontor_resources.classes.parcel import Parcel
from inkontor_resources.classes.price import PriceGls


# todo maybe this check should be included in the parcel class better?
def check_weight(value, parcel):
    if isinstance(value, str):
        raise ValueError(f'weight cant be a string for parcel {parcel.sku}')
    if value == 0:
        raise ValueError(f'weight cant be 0 for parcel {parcel.sku}')
    else:
        return value


# Check if all sub parcels have a corresponding product if not it will raise value error.
def check_sub_parcels(project_path, parcels_file, parcels_workbook):
    os.chdir(project_path)
    sku_workbook = openpyxl.load_workbook(parcels_file)
    sku_sheet = sku_workbook[parcels_workbook]
    products = []
    # read the whole first column in products array because we need it later to check if sub parcels are in
    # products array
    for i in range(3, sku_sheet.max_row + 1):
        if sku_sheet.cell(row=i, column=1).value not in products:
            products.append(sku_sheet.cell(row=i, column=1).value)
    # here we create the sub parcel array
    parcels = get_parcels_sku(project_path, parcels_file, parcels_workbook)
    # check if the sub parcels have a corresponding value in the product array. If not we raise an ValueError
    for parcel in parcels:
        if parcel not in products:
            raise ValueError(f'{parcel} has no values in source file')


# get all parcel sku from the file
def get_parcels_sku(project_path, parcels_file, parcels_workbook):
    os.chdir(project_path)
    sku_workbook = openpyxl.load_workbook(parcels_file)
    sku_sheet = sku_workbook[parcels_workbook]
    parcels_sku = []
    for i in range(3, sku_sheet.max_row + 1):
        for j in range(12, 18):
            # check and filter out unusable characters
            if sku_sheet.cell(row=i, column=j).value is not None and sku_sheet.cell(row=i, column=j).value != '-':
                # add sku to array and check for duplicates
                if sku_sheet.cell(row=i, column=j).value not in parcels_sku:
                    parcels_sku.append(sku_sheet.cell(row=i, column=j).value)
    return parcels_sku


# Read in all products into objects from the source file and create a products array with all product objects.
def read_in_all_products(project_path, parcels_file, parcels_workbook):
    os.chdir(project_path)
    sku_workbook = openpyxl.load_workbook(parcels_file)
    sku_sheet = sku_workbook[parcels_workbook]
    products = []
    parcels_sku = get_parcels_sku(project_path, parcels_file, parcels_workbook)
    # creates the products array with Products objects
    for i in range(3, sku_sheet.max_row + 1):
        # filter out sub parcel sku
        if sku_sheet.cell(row=i, column=1).value not in parcels_sku:
            products.append(Product(sku_sheet.cell(row=i, column=1).value, sku_sheet.cell(row=i, column=2).value))
    return products


# Read in all parcels to the parcel objects and create a parcels array with all parcels
def read_in_all_parcels(project_path, parcels_file, parcels_workbook):
    os.chdir(project_path)
    sku_workbook = openpyxl.load_workbook(parcels_file)
    sku_sheet = sku_workbook[parcels_workbook]
    # get all parcel sku from the file
    parcels_sku = get_parcels_sku(project_path, parcels_file, parcels_workbook)
    parcels = []
    # Iterate through the file and if a sub parcel sku matches with the value in the parcels_sku array we create an
    # instance of the Parcel object.
    for i in range(3, sku_sheet.max_row + 1):
        if sku_sheet.cell(row=i, column=1).value in parcels_sku:
            parcels.append(Parcel(sku_sheet.cell(row=i, column=1).value, sku_sheet.cell(row=i, column=2).value))
    return parcels


def add_parcels_to_product(project_path, parcels_file, parcels_workbook, products, parcels):
    os.chdir(project_path)
    sku_workbook = openpyxl.load_workbook(parcels_file)
    sku_sheet = sku_workbook[parcels_workbook]
    # go through file line by line
    for i in range(3, sku_sheet.max_row + 1):
        # grab the product
        for product in products:
            if product.sku == sku_sheet.cell(row=i, column=1).value:
                for j in range(12, 18):
                    if sku_sheet.cell(row=i, column=j).value is not None \
                            and sku_sheet.cell(row=i, column=j).value != '-':
                        # grab parcel
                        for parcel in parcels:
                            if parcel.sku == sku_sheet.cell(row=i, column=j).value:
                                product.add_parcel(parcel, product)


def add_weight(project_path, parcels_file, parcels_workbook, products, parcels):
    os.chdir(project_path)
    sku_workbook = openpyxl.load_workbook(parcels_file)
    sku_sheet = sku_workbook[parcels_workbook]
    # Add weights to products without sub parcels
    for i in range(3, sku_sheet.max_row + 1):
        # Iterate through all products and filter the product for the sku. If product has no sub-parcels, set weight
        for product in products:
            if product.sku == sku_sheet.cell(row=i, column=1).value:
                if len(product.parcels) == 0:
                    product.weight = sku_sheet.cell(row=i, column=10).value
        for parcel in parcels:
            if parcel.sku == sku_sheet.cell(row=i, column=1).value:
                parcel.set_weight(sku_sheet.cell(row=i, column=10).value)


def create_prices(project_path, price_file_gls, price_workbook_gls):
    os.chdir(project_path)
    price_workbook = openpyxl.load_workbook(price_file_gls)
    price_sheet = price_workbook[price_workbook_gls]
    prices = []
    for i in range(2, price_sheet.max_row + 1):
        prices.append(PriceGls(price_sheet.cell(row=i, column=1).value, price_sheet.cell(row=i, column=2).value,
                               price_sheet.cell(row=i, column=3).value, price_sheet.cell(row=i, column=4).value,
                               price_sheet.cell(row=i, column=5).value, price_sheet.cell(row=i, column=6).value,
                               price_sheet.cell(row=i, column=7).value, price_sheet.cell(row=i, column=8).value))
    return prices


def set_out_of_range(products, parcels):
    for product in products:
        if product.weight > 40 and len(product.parcels) == 0:
            product.weight_out_of_range = True
    for parcel in parcels:
        if parcel.weight > 40:
            parcel.update_out_of_range()
    return False


def set_product_prices(products, prices):
    for product in products:
        shipment_price_product = {}
        if len(product.parcels) == 0:
            if product.weight <= 5:
                for price in prices:
                    shipment_price_product[price.country] = price.to05kg
            if 5 < product.weight <= 10:
                for price in prices:
                    shipment_price_product[price.country] = price.to10kg
            if 10 < product.weight <= 15:
                for price in prices:
                    shipment_price_product[price.country] = price.to15kg
            if 15 < product.weight <= 25:
                for price in prices:
                    shipment_price_product[price.country] = price.to25kg
            if 25 < product.weight <= 30:
                for price in prices:
                    shipment_price_product[price.country] = price.to30kg
            if 30 < product.weight <= 40:
                for price in prices:
                    shipment_price_product[price.country] = price.to40kg
            if product.weight > 40:
                for price in prices:
                    shipment_price_product[price.country] = 0
            product.prices = shipment_price_product


def set_parcel_prices(parcels, prices):
    for parcel in parcels:
        shipment_price_parcel = {}
        if parcel.weight <= 5:
            for price in prices:
                shipment_price_parcel[price.country] = price.to05kg
        if 5 < parcel.weight <= 10:
            for price in prices:
                shipment_price_parcel[price.country] = price.to10kg
        if 10 < parcel.weight <= 15:
            for price in prices:
                shipment_price_parcel[price.country] = price.to15kg
        if 15 < parcel.weight <= 25:
            for price in prices:
                shipment_price_parcel[price.country] = price.to25kg
        if 25 < parcel.weight <= 30:
            for price in prices:
                shipment_price_parcel[price.country] = price.to30kg
        if 30 < parcel.weight <= 40:
            for price in prices:
                shipment_price_parcel[price.country] = price.to40kg
        if parcel.weight > 40:
            for price in prices:
                shipment_price_parcel[price.country] = 0
        parcel.set_price(shipment_price_parcel)
    return parcels


def create_excel(products, sheet, target_file):
    price_list = openpyxl.Workbook()
    price_list.sheetnames
    sheet = price_list.get_sheet_by_name(sheet)
    sheet.cell(row=1, column=1).value = 'Product SKU'
    sheet.cell(row=1, column=2).value = 'Product name'
    sheet.cell(row=1, column=3).value = 'Weight out of range'
    sheet.cell(row=1, column=4).value = 'Price Germany'
    sheet.cell(row=1, column=5).value = 'Price France'
    sheet.cell(row=1, column=6).value = 'Price Italy'
    sheet.cell(row=1, column=7).value = 'Price Spain'
    i = 2
    for product in products:
        sheet.cell(row=i, column=1).value = product.sku
        sheet.cell(row=i, column=2).value = product.name
        sheet.cell(row=i, column=3).value = product.weight_out_of_range
        sheet.cell(row=i, column=4).value = product.prices['Niemcy (6)']
        sheet.cell(row=i, column=5).value = product.prices['Francja I (kontynent)']
        sheet.cell(row=i, column=6).value = product.prices['Włochy']
        sheet.cell(row=i, column=7).value = product.prices['Hiszpania I (kontynent)']
        # for price in product.prices:
        #    if price == 'Niemcy (6)':
        #        sheet.cell(row=i, column=3).value = price[product.prices]
        i = i + 1
    price_list.save(target_file)
    price_list.close()
